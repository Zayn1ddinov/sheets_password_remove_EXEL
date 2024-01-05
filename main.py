from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
import tkinter as tk
from tkinter import filedialog

def open_file_dialog():
    root = tk.Tk()
    root.withdraw()  # Скрыть основное окно
    file_path = filedialog.askopenfilename()  # Открывает окно выбора файла
    return file_path

file_path = open_file_dialog()

# Загрузка Excel-файла
workbook = load_workbook(file_path)

# Получение списка листов в книге
sheets = workbook.sheetnames

for sheet_name in sheets:
    sheet = workbook[sheet_name]
    # Проверка наличия защиты листа
    if sheet.protection.sheet:
        # Установка нового пустого пароля
        sheet.protection = SheetProtection()

# Сохранение изменений
workbook.save('измененный_файл.xlsx')